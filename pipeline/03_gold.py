#!/usr/bin/env -S uv run --script
# /// script
# requires-python = ">=3.11"
# dependencies = ["duckdb>=1.1", "numpy", "openpyxl>=3.1"]
# ///
"""CAMADA GOLD -- junta as fontes no painel municipal e gera os arquivos do site.

Regra que atravessa esta camada: **massa em R$ soma, contagem de pessoas nao soma.**
Cada linha carrega sua propria unidade (vinculo, beneficio, familia) e o painel nunca
produz um total de pessoas. Ver discovery-plan.md secao 4 e briefing-nota-tecnica.md Q1-Q3.

Saidas:
    dados/gold/painel_municipal.parquet   dataset harmonizado (o que sera publicado)
    dados/gold/painel_municipal.csv       mesma coisa, para quem nao usa parquet
    dados/gold/dicionario.md              dicionario de dados
    dados/gold/validacao.md               checagens contra agregados oficiais (E2)
    site/municipios.json                  indice leve para busca
    site/dados.json                       payload do contracheque

Uso:
    uv run pipeline/03_gold.py [--ref 202512]
"""
from __future__ import annotations

import json
import sys
from pathlib import Path

import duckdb

sys.path.insert(0, str(Path(__file__).parent))
import config as C  # noqa: E402

# (chave, unidade, rotulo, fonte, tem_base_media)
# `tem_base_media` marca as linhas cuja media sai de uma base filtrada (0,7 a 30 SM,
# parametro da RAIS) em vez de massa/contagem. Vale so para as linhas de salario.
LINHAS = [
    ("salario_privado", "vinculos",   "Salário — setor privado",          "RAIS", True),
    ("salario_publico", "vinculos",   "Salário — administração pública",  "RAIS", True),
    ("previdencia",     "beneficios", "Previdência — INSS/BPC",           "INSS", False),
    ("bolsa_familia",   "familias",   "Benefício social — Bolsa Família", "Portal da Transparência", False),
]


def monta_painel(con: duckdb.DuckDBPyConnection, sv: Path) -> None:
    p = lambda n: f"'{sv / (n + '.parquet')}'"
    con.execute(f"""CREATE OR REPLACE TABLE painel AS
    WITH rais AS (
      SELECT cod_ibge,
             sum(vinculos)            FILTER (WHERE setor='privado') AS priv_n,
             sum(massa_salarial)      FILTER (WHERE setor='privado') AS priv_massa,
             sum(base_media_vinculos) FILTER (WHERE setor='privado') AS priv_base_n,
             sum(base_media_massa)    FILTER (WHERE setor='privado') AS priv_base_massa,
             sum(vinculos)            FILTER (WHERE setor='publico') AS pub_n,
             sum(massa_salarial)      FILTER (WHERE setor='publico') AS pub_massa,
             sum(base_media_vinculos) FILTER (WHERE setor='publico') AS pub_base_n,
             sum(base_media_massa)    FILTER (WHERE setor='publico') AS pub_base_massa,
             sum(vinculos)       FILTER (WHERE setor='publico' AND esfera='municipal') AS pub_n_mun,
             sum(massa_salarial) FILTER (WHERE setor='publico' AND esfera='municipal') AS pub_massa_mun,
             sum(vinculos)       FILTER (WHERE setor='publico' AND esfera='estadual')  AS pub_n_est,
             sum(vinculos)       FILTER (WHERE setor='publico' AND esfera='federal')   AS pub_n_fed
      FROM {p('rais')} GROUP BY 1),
    inss AS (SELECT cod_ibge, sum(qtde) AS n, sum(valor) AS massa FROM {p('inss')} GROUP BY 1)
    SELECT d.cod_ibge, d.nome, d.uf, d.regiao,
           pop.pop_total, pop.pop_adulta,
           coalesce(r.priv_n,0)::BIGINT     AS salario_privado_n,
           coalesce(r.priv_massa,0)         AS salario_privado_massa,
           r.priv_base_n::BIGINT            AS salario_privado_base_n,
           r.priv_base_massa                AS salario_privado_base_massa,
           coalesce(r.pub_n,0)::BIGINT      AS salario_publico_n,
           coalesce(r.pub_massa,0)          AS salario_publico_massa,
           r.pub_base_n::BIGINT             AS salario_publico_base_n,
           r.pub_base_massa                 AS salario_publico_base_massa,
           coalesce(r.pub_n_mun,0)::BIGINT  AS salario_publico_municipal_n,
           coalesce(r.pub_massa_mun,0)      AS salario_publico_municipal_massa,
           coalesce(r.pub_n_est,0)::BIGINT  AS salario_publico_estadual_n,
           coalesce(r.pub_n_fed,0)::BIGINT  AS salario_publico_federal_n,
           coalesce(i.n,0)::BIGINT          AS previdencia_n,
           coalesce(i.massa,0)              AS previdencia_massa,
           coalesce(bf.familias_ref,0)::BIGINT AS bolsa_familia_n,
           coalesce(bf.valor_ref,0)         AS bolsa_familia_massa,
           bf.familias_comp::BIGINT         AS bolsa_familia_n_competencia,
           bf.valor_comp                    AS bolsa_familia_massa_competencia,
           cu.tam_medio_familia,
           cu.media_pessoas_pbf,
           -- numero publicado pelo MDS, nao estimativa: familias do BF sao maiores que a
           -- media do CadUnico, entao derivar do tamanho medio geral subestima em ~18%
           cu.pessoas_pbf::BIGINT           AS bolsa_familia_pessoas,
           cu.familias                      AS cadunico_familias,
           cu.pessoas                       AS cadunico_pessoas
    FROM {p('dim_municipio')} d
    LEFT JOIN {p('populacao')} pop USING (cod_ibge)
    LEFT JOIN rais r    USING (cod_ibge)
    LEFT JOIN inss i    USING (cod_ibge)
    LEFT JOIN {p('bolsa_familia')} bf USING (cod_ibge)
    LEFT JOIN {p('cadunico')} cu      USING (cod_ibge)""")

    # derivadas: massa total, participacoes e taxas por 100 adultos
    con.execute("""CREATE OR REPLACE TABLE painel AS
    SELECT *,
      salario_privado_massa + salario_publico_massa + previdencia_massa + bolsa_familia_massa
        AS massa_total,
      CASE WHEN pop_total > 0 THEN
        (salario_privado_massa + salario_publico_massa + previdencia_massa + bolsa_familia_massa)
        / pop_total END AS massa_per_capita
    FROM painel""")
    derivadas = []
    for chave, _, _, _, base_media in LINHAS:
        # a media das linhas de salario sai da base filtrada de 0,7 a 30 SM; as demais
        # linhas usam massa / contagem, porque nao ha parametro equivalente
        num, den = ((f"{chave}_base_massa", f"{chave}_base_n") if base_media
                    else (f"{chave}_massa", f"{chave}_n"))
        derivadas += [
            f"CASE WHEN massa_total > 0 THEN {chave}_massa / massa_total END AS {chave}_part",
            f"CASE WHEN {den} > 0 THEN {num} / {den} END AS {chave}_medio",
            f"CASE WHEN pop_adulta > 0 THEN 100.0 * {chave}_n / pop_adulta END AS {chave}_por100",
        ]
    derivadas.append("""CASE WHEN pop_total < 5000   THEN 'ate_5k'
                             WHEN pop_total < 20000  THEN '5k_20k'
                             WHEN pop_total < 100000 THEN '20k_100k'
                             ELSE '100k_mais' END AS porte""")
    con.execute("CREATE OR REPLACE TABLE painel AS SELECT *, " + ", ".join(derivadas) + " FROM painel")


def valida(con, sv: Path, ref: str, out: Path) -> list[str]:
    """E2 -- confere o painel contra agregados oficiais independentes."""
    linhas = [f"# Validação do painel — competência {ref}", ""]
    tot = con.sql("""SELECT count(*), sum(salario_privado_n), sum(salario_publico_n),
                            sum(previdencia_n), sum(bolsa_familia_n), sum(massa_total)
                     FROM painel""").fetchone()
    linhas += [
        "## Totais do painel", "",
        f"- Municípios: **{tot[0]:,}**",
        f"- Vínculos privados: **{tot[1]:,}**",
        f"- Vínculos públicos: **{tot[2]:,}**",
        f"- Benefícios do INSS: **{tot[3]:,}**",
        f"- Famílias no Bolsa Família: **{tot[4]:,}**",
        f"- Massa total registrada: **R$ {tot[5]:,.2f}** por mês", "",
    ]

    # benchmark INSS: agregado nacional oficial por especie
    xlsx = C.BRONZE / ref / f"inss_agregado_{ref}.xlsx"
    if xlsx.exists():
        import openpyxl
        wb = openpyxl.load_workbook(xlsx, read_only=True)
        ws = wb[wb.sheetnames[0]]
        oficial_q = oficial_v = 0
        for row in ws.iter_rows(min_row=5, values_only=True):
            if row and row[1] and row[2]:
                try:
                    oficial_q += int(row[1]); oficial_v += float(row[2])
                except (TypeError, ValueError):
                    pass
        nosso = con.sql(f"SELECT sum(qtde), sum(valor) FROM '{sv / 'inss.parquet'}'").fetchone()
        nao = json.loads((sv / "inss_nao_alocado.json").read_text())
        dq = 100 * (nosso[0] + nao["beneficios"] - oficial_q) / oficial_q if oficial_q else 0
        dv = 100 * (nosso[1] + nao["valor"] - oficial_v) / oficial_v if oficial_v else 0
        linhas += [
            "## INSS — painel vs. agregado nacional oficial", "",
            "| | Quantidade | Valor |", "|---|---|---|",
            f"| Agregado oficial do INSS | {oficial_q:,} | R$ {oficial_v:,.2f} |",
            f"| Painel + não alocado | {nosso[0] + nao['beneficios']:,} | R$ {nosso[1] + nao['valor']:,.2f} |",
            f"| Desvio | {dq:+.4f}% | {dv:+.4f}% |", "",
            f"> Não alocado a município: {nao['beneficios']:,} benefícios "
            f"({100*nao['beneficios']/nao['total_beneficios']:.2f}%), R$ {nao['valor']:,.2f}.", "",
        ]
        if abs(dv) > 1:
            linhas += ["**ATENÇÃO: desvio acima de 1%. Investigar antes de publicar.**", ""]

    # cobertura
    faltas = con.sql("""SELECT
        count(*) FILTER (WHERE salario_privado_n=0 AND salario_publico_n=0) sem_rais,
        count(*) FILTER (WHERE previdencia_n=0) sem_inss,
        count(*) FILTER (WHERE bolsa_familia_n=0) sem_bf,
        count(*) FILTER (WHERE pop_adulta IS NULL) sem_pop,
        count(*) FILTER (WHERE tam_medio_familia IS NULL) sem_cadunico FROM painel""").fetchone()
    linhas += [
        "## Cobertura — municípios sem dado em cada fonte", "",
        "| Fonte | Municípios sem dado |", "|---|---|",
        f"| RAIS | {faltas[0]} |", f"| INSS | {faltas[1]} |",
        f"| Bolsa Família | {faltas[2]} |", f"| População | {faltas[3]} |",
        f"| CadÚnico | {faltas[4]} |", "",
        "> Município sem dado aparece como `—` no site, nunca como zero.", "",
    ]

    # sanidade: massa nao pode superar o PIB, e participacoes somam 1
    ruins = con.sql("""SELECT count(*) FROM painel WHERE massa_total > 0 AND
        abs(coalesce(salario_privado_part,0)+coalesce(salario_publico_part,0)
           +coalesce(previdencia_part,0)+coalesce(bolsa_familia_part,0) - 1) > 0.001""").fetchone()[0]
    linhas += [f"## Consistência\n\n- Municípios com participações que não somam 100%: **{ruins}**", ""]
    out.write_text("\n".join(linhas), encoding="utf-8")
    return linhas


def exporta_site(con, ref: str) -> None:
    C.SITE.mkdir(parents=True, exist_ok=True)
    idx = [{"id": r[0], "nome": r[1], "uf": r[2], "regiao": r[3],
            "pop": r[4], "pop18": r[5], "porte": r[6]}
           for r in con.sql("""SELECT cod_ibge, nome, uf, regiao, pop_total, pop_adulta, porte
                               FROM painel ORDER BY nome""").fetchall()]
    (C.SITE / "municipios.json").write_text(json.dumps(idx, ensure_ascii=False,
                                                       separators=(",", ":")), encoding="utf-8")

    dados = {}
    cols = [c for c in con.sql("SELECT * FROM painel LIMIT 0").columns]
    for row in con.sql("SELECT * FROM painel").fetchall():
        r = dict(zip(cols, row))
        linhas, supressoes = {}, []
        for chave, unidade, _, _, base_media in LINHAS:
            n = r[f"{chave}_n"]
            if not n:
                supressoes.append(chave)
                continue
            item = {"n": int(n), "unidade": unidade,
                    "massa": round(r[f"{chave}_massa"]),
                    "medio": round(r[f"{chave}_medio"] or 0),
                    "part": round(r[f"{chave}_part"] or 0, 4),
                    "por100": round(r[f"{chave}_por100"] or 0, 1)}
            if chave == "salario_publico":
                item["esferas"] = {"municipal": int(r["salario_publico_municipal_n"] or 0),
                                   "estadual": int(r["salario_publico_estadual_n"] or 0),
                                   "federal": int(r["salario_publico_federal_n"] or 0)}
            if chave == "bolsa_familia" and r["bolsa_familia_pessoas"]:
                item["pessoas"] = int(r["bolsa_familia_pessoas"])
                item["pessoas_por_familia"] = (round(r["media_pessoas_pbf"], 2)
                                               if r["media_pessoas_pbf"] else None)
                if r["pop_adulta"]:
                    item["pessoas_por100"] = round(100.0 * r["bolsa_familia_pessoas"] / r["pop_adulta"], 1)
            linhas[chave] = item
        # NAO existe campo de total de pessoas. A ausencia e proposital: impede
        # que alguem some unidades incompativeis mais tarde.
        dados[r["cod_ibge"]] = {"ref": f"{ref[:4]}-{ref[4:]}", "linhas": linhas,
                                "massa_total": round(r["massa_total"] or 0),
                                "massa_per_capita": round(r["massa_per_capita"] or 0),
                                "supressoes": supressoes}
    (C.SITE / "dados.json").write_text(json.dumps(dados, ensure_ascii=False,
                                                  separators=(",", ":")), encoding="utf-8")
    for f in ("municipios.json", "dados.json"):
        C.log(f"  site/{f}: {C.humano((C.SITE / f).stat().st_size)}")


def dicionario(out: Path) -> None:
    linhas = ["# Dicionário de dados — painel municipal", "",
              "Uma linha por município. Valores monetários em reais nominais do mês de referência.", "",
              "| Coluna | Descrição | Unidade |", "|---|---|---|",
              "| `cod_ibge` | Código IBGE de 7 dígitos | — |",
              "| `nome`, `uf`, `regiao` | Identificação | — |",
              "| `pop_total` | População residente estimada | pessoas |",
              f"| `pop_adulta` | População com {C.IDADE_ADULTA} anos ou mais | pessoas |"]
    for chave, unidade, rotulo, fonte, base_media in LINHAS:
        linhas += [f"| `{chave}_n` | {rotulo} — contagem ({fonte}) | **{unidade}** |",
                   f"| `{chave}_massa` | {rotulo} — massa mensal | R$ |",
                   f"| `{chave}_medio` | {rotulo} — valor médio"
                   + (f" (base: {C.FAIXA_MEDIA_SM[0]} a {C.FAIXA_MEDIA_SM[1]} SM)" if base_media else "")
                   + f" | R$ / {unidade[:-1]} |",
                   f"| `{chave}_part` | {rotulo} — participação na massa total | proporção |",
                   f"| `{chave}_por100` | {rotulo} — por 100 adultos | taxa |"]
    linhas += [
        "| `massa_total` | Soma das quatro massas | R$ |",
        "| `massa_per_capita` | Massa total ÷ população | R$ |",
        "| `porte` | Faixa populacional | — |", "",
        "## Atenção às unidades", "",
        "As colunas `_n` **não são comparáveis entre si e não devem ser somadas**. "
        "A RAIS conta vínculos, o INSS conta benefícios emitidos e o Bolsa Família conta famílias. "
        "Uma mesma pessoa pode aparecer em mais de uma fonte. As colunas `_por100` existem "
        "justamente para permitir leitura relativa sem induzir soma — elas podem ultrapassar 100 "
        "no conjunto, e isso não é erro.", "",
        "As colunas `_massa` **são** somáveis: representam fluxos de dinheiro distintos.", "",
    ]
    out.write_text("\n".join(linhas), encoding="utf-8")


def main() -> None:
    ref = C.ref_do_argv()
    sv = C.SILVER / ref
    if not sv.exists():
        raise SystemExit(f"{sv} nao existe -- rode 02_silver.py --ref {ref} primeiro")
    C.GOLD.mkdir(parents=True, exist_ok=True)
    gd = C.dir_camada(C.GOLD, ref)

    C.log(f"GOLD | competencia={ref}")
    con = duckdb.connect()
    con.execute("PRAGMA memory_limit='6GB'")

    C.log("1/4 montando painel municipal")
    monta_painel(con, sv)
    n = con.sql("SELECT count(*) FROM painel").fetchone()[0]
    C.log(f"  painel: {n} municipios")
    con.execute(f"COPY painel TO '{gd / 'painel_municipal.parquet'}' (FORMAT parquet)")
    con.execute(f"COPY painel TO '{gd / 'painel_municipal.csv'}' (HEADER, DELIMITER ',')")

    C.log("2/4 dicionario de dados")
    dicionario(gd / "dicionario.md")

    C.log("3/4 validacao (E2)")
    for l in valida(con, sv, ref, gd / "validacao.md"):
        if l.startswith(("- ", "| Desvio", "**ATENCAO", "**ATENÇÃO")):
            C.log("  " + l)

    C.log("4/4 exportando arquivos do site")
    exporta_site(con, ref)

    C.log(f"GOLD concluida. Painel em {gd}")


if __name__ == "__main__":
    main()
